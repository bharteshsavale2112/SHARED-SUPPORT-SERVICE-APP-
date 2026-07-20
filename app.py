from flask import Flask, request, jsonify, send_from_directory, session
from functools import wraps
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from flask import request, jsonify
import openpyxl
import pandas as pd
import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


app = Flask(__name__)

# ==========================
# SECURITY CONFIG
# ==========================
# Secret key used to sign session cookies. Set SECRET_KEY as an
# environment variable on Render (Settings -> Environment) in production.
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")

# ---- Cookie / Session hardening ----
# HTTPONLY: JavaScript can't read the session cookie (protects against XSS
#           stealing the login session)
# SAMESITE: cookie is not sent on cross-site requests (protects against CSRF)
# SECURE:   cookie is only sent over HTTPS. Render serves the site over
#           HTTPS, so this is safe to enable in production. If you ever
#           test locally over plain http://127.0.0.1, set
#           FLASK_DEBUG=True (which also flips this off automatically).
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("FLASK_DEBUG", "False") != "True"

# Auto logout after 2 hours of inactivity
from datetime import timedelta as _timedelta
app.config["PERMANENT_SESSION_LIFETIME"] = _timedelta(hours=2)

# Admin credentials now come from environment variables instead of being
# hardcoded in the source code (which is visible to anyone with repo access).
# Set ADMIN_EMAIL and ADMIN_PASSWORD as environment variables on Render.
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@bajaj.com")
ADMIN_PASSWORD_HASH = os.environ.get("ADMIN_PASSWORD_HASH") or generate_password_hash(
    os.environ.get("ADMIN_PASSWORD", "admin123")
)


def login_required(role=None):
    """Blocks access to a route unless the user is logged in
    (and, if `role` is given, logged in as that specific role)."""
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if "role" not in session:
                return jsonify({"status": "error", "message": "Login Required"}), 401
            if role and session.get("role") != role:
                return jsonify({"status": "error", "message": "Not Authorized"}), 403
            return f(*args, **kwargs)
        return wrapper
    return decorator


EMPLOYEE_FILE = "employees.xlsx"

# ==========================
# CREATE EXCEL FILE SAFELY
# ==========================

def create_excel_if_needed():

    cols = [
    "employeeCode",
    "employeeName",
    "mobile",
    "department",
    "email",
    "shift",
    "busStop",
    "routeNumber",
    "address",
    "password",
    "route"
]

    if not os.path.exists(EMPLOYEE_FILE):
        pd.DataFrame(columns=cols).to_excel(EMPLOYEE_FILE, index=False)
        return

    try:
        df = pd.read_excel(EMPLOYEE_FILE)
        df.columns = df.columns.str.strip()

        for col in cols:
            if col not in df.columns:
                df[col] = ""

        df = df[cols]
        df.to_excel(EMPLOYEE_FILE, index=False)

    except Exception:
        pd.DataFrame(columns=cols).to_excel(EMPLOYEE_FILE, index=False)

create_excel_if_needed()



# ==========================
# LOAD EXCEL SAFELY
# ==========================

def load_df():
    create_excel_if_needed()   # <-- ही नवीन line add कर

    df = pd.read_excel(EMPLOYEE_FILE)
    df.columns = df.columns.str.strip()

    return df

# ==========================
# HOME
# ==========================

@app.route("/")
def home():
    return """
    <h2>🚍 Bajaj Transport Backend Running</h2>
    <a href="/employeeregistration.html">Employee Registration</a><br>
    <a href="/employeelogin.html">Employee Login</a><br>
    <a href="/adminlogin.html">Admin Login</a>
    """

# ==========================
# SERVE HTML FILES
# ==========================

@app.route("/<path:path>")
def serve(path):
    return send_from_directory(".", path)








# ==========================
# EMPLOYEE REGISTRATION
# ==========================

@app.route("/employee-registration", methods=["POST"])
def register():

    try:

        df = load_df()

        data = request.form.to_dict()

        # Required Fields
        required = [
            "employeeCode",
            "employeeName",
            "mobile",
            "department",
            "email",
            "shift",
            "busStop",
            "routeNumber",
            "address",
            "password"
        ]

        for field in required:

            if field not in data or str(data[field]).strip() == "":
                return f"<script>alert('{field} is Required');history.back()</script>"

        # Duplicate Employee Code
        if str(data["employeeCode"]).strip() in df["employeeCode"].astype(str).str.strip().values:

            return "<script>alert('Employee Code Already Registered');history.back()</script>"

        # Duplicate Email
        if str(data["email"]).strip().lower() in df["email"].astype(str).str.strip().str.lower().values:

            return "<script>alert('Email Already Registered');history.back()</script>"

        # Duplicate Mobile
        if str(data["mobile"]).strip() in df["mobile"].astype(str).str.strip().values:

            return "<script>alert('Mobile Number Already Registered');history.back()</script>"

        # Default Route
        data["route"] = ""

        # Hash the password before saving so it's never stored in plain text
        data["password"] = generate_password_hash(str(data["password"]))

        # Save Employee
        df = pd.concat(
            [df, pd.DataFrame([data])],
            ignore_index=True
        )

        df.to_excel(EMPLOYEE_FILE, index=False)

        return """
        <script>
        alert('Registration Successful');
        window.location.href='/employeelogin.html';
        </script>
        """

    except Exception as e:

        print(e)

        return f"""
        <script>
        alert('{str(e)}');
        history.back();
        </script>
        """
# ==========================
# LOGIN
# ==========================

@app.route("/employee-login", methods=["POST"])
def employee_login():

    try:

        data = request.get_json()

        if not data:
            return jsonify({
                "status": "error",
                "message": "Invalid Request"
            })

        employee_id = str(data.get("employeeId", "")).strip()
        password = str(data.get("password", "")).strip()

        if employee_id == "" or password == "":
            return jsonify({
                "status": "error",
                "message": "Employee ID and Password Required"
            })

        # Load Excel Data
        df = load_df()

        # Find Employee
        user = df[
            df["employeeCode"].astype(str).str.strip() == employee_id
        ]

        if user.empty:
            return jsonify({
                "status": "error",
                "message": "Employee ID Not Found"
            })

        # Employee Record
        row = user.iloc[0]

        # Password Check (supports old plain-text passwords too, and
        # upgrades them to a secure hash automatically on next login)
        db_password = str(row["password"]).strip()

        if db_password.startswith(("pbkdf2:", "scrypt:")):
            password_ok = check_password_hash(db_password, password)
        else:
            password_ok = (db_password == password)
            if password_ok:
                # Legacy plain-text password found; upgrade it to a hash now
                df.loc[user.index, "password"] = generate_password_hash(password)
                df.to_excel(EMPLOYEE_FILE, index=False)

        if not password_ok:
            return jsonify({
                "status": "error",
                "message": "Wrong Password"
            })

        # Login Success - mark this browser session as logged in
        session.permanent = True
        session["role"] = "employee"
        session["employeeCode"] = employee_id

        return jsonify({

            "status": "success",
            "message": "Login Successful",

            "passNumber": str(row.get("passNumber", "")),
            "employeeCode": str(row.get("employeeCode", "")).strip(),
            "employeeName": str(row.get("employeeName", "")),
            "department": str(row.get("department", "")),
            "mobile": str(row.get("mobile", "")),
            "shift": str(row.get("shift", "")),
            "routeNumber": str(row.get("routeNumber", "")),
            "busStop": str(row.get("busStop", "")),
            "issueDate": str(row.get("issueDate", "")),
            "passStatus": str(row.get("status", ""))

        })

    except Exception as e:

        return jsonify({
            "status": "error",
            "message": str(e)
        })
            
# ==========================
# ADMIN LOGIN
# ==========================

@app.route("/admin-login", methods=["POST"])
def admin_login():

    data = request.get_json()

    email = str(data.get("email", "")).strip()
    password = str(data.get("password", "")).strip()

    if email == ADMIN_EMAIL and check_password_hash(ADMIN_PASSWORD_HASH, password):
        session.permanent = True
        session["role"] = "admin"
        return jsonify({
            "status": "success",
            "message": "Login Successful"
        })

    return jsonify({
        "status": "error",
        "message": "Wrong Credentials"
    })


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"status": "success", "message": "Logged Out"})


@app.route("/check-session", methods=["GET"])
def check_session():
    """Used by pages to verify the login session is still valid
    (e.g. after using the browser Back/Forward buttons post-logout)."""
    if "role" not in session:
        return jsonify({"status": "error", "message": "Not Logged In"}), 401
    return jsonify({
        "status": "success",
        "role": session.get("role"),
        "employeeCode": session.get("employeeCode", "")
    })


@app.after_request
def add_no_cache_headers(response):
    """Stops the browser from serving a cached copy of protected pages
    from its Back/Forward cache after the user has logged out."""
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response





# ==========================
# ADMIN DASHBOARD STATS
# ==========================

@app.route("/dashboard-stats")
@login_required("admin")
def dashboard_stats():

    try:

        emp = load_df()

        totalEmployees = len(emp)

        totalRoutes = emp["route"].replace("", pd.NA).dropna().nunique()

        pendingRequests = 0
        activePasses = 0

        if os.path.exists(PASS_REQUEST_FILE):

            req = pd.read_excel(PASS_REQUEST_FILE)

            pendingRequests = len(
                req[req["status"] == "Pending"]
            )

        if os.path.exists(PASS_FILE):

            passes = pd.read_excel(PASS_FILE)

            activePasses = len(passes)

        return jsonify({

            "totalEmployees": totalEmployees,

            "totalRoutes": totalRoutes,

            "pendingRequests": pendingRequests,

            "activePasses": activePasses

        })

    except Exception as e:

        return jsonify({

            "status": "error",

            "message": str(e)

        })
    

# ==========================
# RECENT EMPLOYEES API
# ==========================

@app.route("/recent-employees")
@login_required("admin")
def recent_employees():

    try:

        df = load_df()

        # शेवटचे 5 Employees
        df = df.tail(5)

        employees = []

        for _, row in df.iterrows():

            employees.append({

                "employeeCode": str(row["employeeCode"]),

                "employeeName": str(row["employeeName"]),

                "department": str(row["department"]),

                "route": str(row["route"])

            })

        return jsonify(employees)

    except Exception as e:

        return jsonify({
            "status":"error",
            "message":str(e)
        })


# ==========================
# VIEW EMPLOYEES
# ==========================

@app.route("/employees")
@login_required("admin")
def employees():

    try:

        df = load_df()

        if "password" in df.columns:
            df = df.drop(columns=["password"])

        return df.to_html(index=False)

    except Exception as e:

        return f"<h3>Error : {str(e)}</h3>"
    

# ==========================
# ASSIGN ROUTE
# ==========================

@app.route("/assign-route", methods=["POST"])
@login_required("admin")
def assign():

    try:

        df = load_df()

        code = str(request.form.get("employeeCode", "")).strip()
        route = str(request.form.get("route", "")).strip()

        # Empty Validation
        if code == "" or route == "":
            return jsonify({
                "success": False,
                "message": "Employee Code and Route Required"
            })

        # Employee Exists?
        if code not in df["employeeCode"].astype(str).str.strip().values:
            return jsonify({
                "success": False,
                "message": "Employee Not Found"
            })

        # Assign Route
        df.loc[
            df["employeeCode"].astype(str).str.strip() == code,
            "route"
        ] = route

        df.to_excel(EMPLOYEE_FILE, index=False)

        return jsonify({
            "success": True,
            "message": "Route Assigned Successfully"
        })

    except Exception as e:

        return jsonify({
            "success": False,
            "message": str(e)
        })
    
# ==========================
# GPS STORAGE (LIVE TRACKING)
# ==========================

live = {}

@app.route("/update-location", methods=["POST"])
def update():

    try:

        data = request.get_json()

        if (
            not data or
            "route" not in data or
            "lat" not in data or
            "lng" not in data
        ):
            return jsonify({
                "success": False,
                "message": "Invalid Data"
            })

        route = str(data["route"]).strip()

        live[route] = {
            "lat": float(data["lat"]),
            "lng": float(data["lng"])
        }

        return jsonify({
            "success": True,
            "message": "Location Updated"
        })

    except Exception as e:

        return jsonify({
            "success": False,
            "message": str(e)
        })

#=========================
# live location route
#=========================

@app.route("/live-location/<route>")
def get(route):

    route = str(route).strip()

    if route not in live:
        return jsonify({
            "success": False,
            "message": "Location Not Found"
        })

    return jsonify({
        "success": True,
        "lat": live[route]["lat"],
        "lng": live[route]["lng"]
    })







# ==========================
# PASS FILE
# ==========================

from datetime import datetime

PASS_FILE = "passes.xlsx"

# ==========================
# CREATE PASSES EXCEL
# ==========================

def create_pass_file():

    cols=[

"passNumber",

"employeeCode",

"employeeName",

"department",

"mobile",

"shift",

"routeNumber",

"busStop",

"issueDate",

"status"

]
    if os.path.exists(PASS_FILE):

        try:

            df = pd.read_excel(PASS_FILE)

            df.columns = df.columns.str.strip()

            for col in cols:

                if col not in df.columns:

                    df[col] = ""

            df = df[cols]

            df.to_excel(
                PASS_FILE,
                index=False
            )

        except Exception:

            pd.DataFrame(columns=cols).to_excel(
                PASS_FILE,
                index=False
            )

    else:

        pd.DataFrame(columns=cols).to_excel(
            PASS_FILE,
            index=False
        )

create_pass_file()

# ==========================
# BUS PASS API
# ==========================

@app.route("/bus-pass", methods=["POST"])
@login_required("employee")
def bus_pass():

    try:

        data = request.get_json()

        # Always use the logged-in session's own employeeCode, never a
        # client-submitted value - fixes false "Not Authorized" and closes
        # off a user tampering with the request to act as someone else.
        employeeCode = session.get("employeeCode")

        df = load_df()

        user = df[
            df["employeeCode"].astype(str) == str(employeeCode)
        ]

        if user.empty:

            return jsonify({
                "status":"error",
                "message":"Employee Not Found"
            })

        pass_df = pd.read_excel(PASS_FILE)

        month = datetime.now().strftime("%B")
        year = datetime.now().year

        already = pass_df[
            (pass_df["employeeCode"].astype(str) == str(employeeCode)) &
            (pass_df["month"] == month) &
            (pass_df["year"] == year)
        ]

        if not already.empty:

            return jsonify({
                "status":"success",
                "message":"Pass Already Generated"
            })

        new_pass = {
            "employeeCode": user.iloc[0]["employeeCode"],
            "employeeName": user.iloc[0]["employeeName"],
            "route": user.iloc[0]["route"],
            "busStop": user.iloc[0]["busStop"],
            "month": month,
            "year": year,
            "created_at": datetime.now().strftime("%d-%m-%Y %H:%M")
        }

        pass_df = pd.concat(
            [pass_df, pd.DataFrame([new_pass])],
            ignore_index=True
        )

        pass_df.to_excel(PASS_FILE, index=False)

        return jsonify({
            "status":"success",
            "message":"Bus Pass Generated Successfully"
        })

    except Exception as e:

        return jsonify({
            "status":"error",
            "message":str(e)
        })

# ==========================
# CHECK PASS API
# ==========================


@app.route("/check-pass", defaults={"employeeCode": None})
@app.route("/check-pass/<employeeCode>")
@login_required("employee")
def check_pass(employeeCode):

    employeeCode = session.get("employeeCode")

    df = pd.read_excel(PASS_FILE)

    user = df[
        df["employeeCode"].astype(str) == str(employeeCode)
    ]

    if user.empty:

        return jsonify({
            "hasPass": False
        })

    return jsonify({

        "hasPass": True,

        "pass": user.iloc[0].to_dict()

    })


# ==========================
#pp request
# ==========================


import uuid
PASS_REQUEST_FILE = "pass_requests.xlsx"

# ==========================
# CREATE PASS REQUEST FILE
# ==========================

# ==========================
# APPROVE PASS
# ==========================

@app.route("/approve-pass", methods=["POST"])
@login_required("admin")
def approve_pass():

    try:
        data = request.get_json()
        requestId = str(data.get("requestId", "")).strip()

        # Load request file
        req = pd.read_excel(PASS_REQUEST_FILE)
        req = req.fillna("")

        req["requestId"] = req["requestId"].astype(str)

        row = req[req["requestId"] == requestId]

        if row.empty:
            return jsonify({
                "status": "error",
                "message": "Request Not Found"
            })

        index = row.index[0]

        # Update request
        req.loc[index, "status"] = "Approved"
        req.loc[index, "approvedBy"] = "Admin"
        req.loc[index, "approvedDate"] = datetime.now().strftime("%d-%m-%Y %H:%M")

        req.to_excel(PASS_REQUEST_FILE, index=False)

        # Load pass file
        passes = pd.read_excel(PASS_FILE)
        passes = passes.fillna("")

        # FIXED COLUMN NAME (IMPORTANT)
        newPass = {

    "passNumber":"PP"+datetime.now().strftime("%Y%m%d%H%M%S"),

    "employeeCode":row.iloc[0]["employeeCode"],

    "employeeName":row.iloc[0]["employeeName"],

    "department":row.iloc[0]["department"],

    "mobile":row.iloc[0]["mobile"],

    "shift":row.iloc[0]["shift"],

    "routeNumber":row.iloc[0]["routeNumber"],

    "busStop":row.iloc[0]["busStop"],

    "issueDate":datetime.now().strftime("%d-%m-%Y"),

    "status":"ACTIVE"

}
        passes = pd.concat([passes, pd.DataFrame([newPass])], ignore_index=True)
        passes.to_excel(PASS_FILE, index=False)

        return jsonify({
            "status": "success",
            "message": "Pass Generated Successfully"
        })

    except Exception as e:

        print("APPROVE PASS ERROR:", e)

        return jsonify({

            "status":"error",

            "message":str(e)

        })

 # ==========================
# GET MY PASS
# ==========================
@app.route("/my-pass", defaults={"employeeCode": None}, methods=["GET"])
@app.route("/my-pass/<employeeCode>", methods=["GET"])
@login_required("employee")
def my_pass(employeeCode):

    # Always trust the SERVER-SIDE session for identity, never the
    # URL/client value. Previously this route 403'd with "Not Authorized"
    # whenever the URL's employeeCode didn't exactly match the session's
    # (e.g. stale localStorage from an earlier login/browser session) —
    # even though the person WAS validly logged in. Since this route is
    # already locked to "your own pass only", the URL value is redundant;
    # using the session value directly removes that whole bug class.
    employeeCode = session.get("employeeCode")

    try:

        print("Employee Code from session:", employeeCode)

        df = pd.read_excel(PASS_FILE)
        df.columns = df.columns.str.strip()

        print(df[["employeeCode"]])

        user = df[
            df["employeeCode"].astype(str).str.strip() ==
            str(employeeCode).strip()
        ]

        print("Matched Rows:", len(user))

        if user.empty:
            return jsonify({
                "status": "error",
                "message": "Pass Not Found"
            })

        row = user.iloc[0]

        return jsonify({
            "status": "success",
            "employeeCode": str(row["employeeCode"]),
            "employeeName": str(row["employeeName"]),
            "department": str(row["department"]),
            "mobile": str(row["mobile"]),
            "shift": str(row["shift"]),
            "routeNumber": str(row["routeNumber"]),
            "route": str(row["routeNumber"]),
            "busStop": str(row["busStop"]),
            "issueDate": str(row["issueDate"]),
            "passStatus": str(row["status"])
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        })
        
# ==========================
# CREATE PASS FILE
# ==========================

def create_pass_file():

    if not os.path.exists(PASS_FILE):

        pd.DataFrame(columns=[

            "passNumber",

            "employeeCode",

            "employeeName",

            "department",

            "mobile",

            "shift",

            "routeNumber",

            "busStop",

            "issueDate",

            "status"

        ]).to_excel(
            PASS_FILE,
            index=False
        )


create_pass_file()




# ==========================
# REQUEST PERMANENT PASS
# ==========================

@app.route("/request-pass", methods=["POST"])
@login_required("employee")
def request_pass():

    try:

        data = request.get_json()

        # Always use the logged-in session's own employeeCode, never a
        # client-submitted value.
        employeeCode = str(session.get("employeeCode", "")).strip()

        if employeeCode == "":

            return jsonify({
                "status": "error",
                "message": "Employee Code Required"
            })

        # Employee Data
        df = load_df()

        user = df[
            df["employeeCode"].astype(str).str.strip()
            == employeeCode
        ]

        if user.empty:

            return jsonify({

                "status": "error",

                "message": "Employee Not Found"

            })

        # Pass Request File
        req = pd.read_excel(PASS_REQUEST_FILE)
        req.columns = req.columns.str.strip()

        # Check Pending Request
        old = req[

            (req["employeeCode"].astype(str).str.strip() == employeeCode)

            &

            (req["status"].astype(str).str.strip() == "Pending")

        ]

        if not old.empty:

            return jsonify({

                "status": "error",

                "message": "Request Already Pending"

            })

        # Generate Request ID
        requestId = "REQ" + datetime.now().strftime("%Y%m%d%H%M%S")

        # New Request
        newRow = {

            "requestId": requestId,

            "employeeCode": str(user.iloc[0]["employeeCode"]),

            "employeeName": str(user.iloc[0]["employeeName"]),

            "department": str(user.iloc[0]["department"]),

            "mobile": str(user.iloc[0]["mobile"]),

            "email": str(user.iloc[0]["email"]),

            "shift": str(user.iloc[0]["shift"]),

            "routeNumber": str(user.iloc[0]["routeNumber"]),

            "busStop": str(user.iloc[0]["busStop"]),

            "requestDate": datetime.now().strftime("%d-%m-%Y %H:%M"),

            "status": "Pending",

            "approvedBy": "",

            "approvedDate": ""

        }

        req = pd.concat(

            [req, pd.DataFrame([newRow])],

            ignore_index=True

        )

        req.to_excel(

            PASS_REQUEST_FILE,

            index=False

        )

        return jsonify({

            "status": "success",

            "message": "Permanent Pass Request Submitted Successfully"

        })

    except Exception as e:

        return jsonify({

            "status": "error",

            "message": str(e)

        })


# ==========================
# PENDING PASS REQUESTS API
# ==========================

@app.route("/pending-pass-requests")
@login_required("admin")
def pending_pass_requests():

    try:

        if not os.path.exists(PASS_REQUEST_FILE):

            return jsonify([])

        df = pd.read_excel(PASS_REQUEST_FILE)

        df = df[df["status"] == "Pending"]

        requests = []

        for _, row in df.iterrows():

            requests.append({

                "requestId": str(row["requestId"]),

                "employeeCode": str(row["employeeCode"]),

                "employeeName": str(row["employeeName"]),

                "route": str(row["routeNumber"]),

                "requestDate": str(row["requestDate"])

            })

        return jsonify(requests)

    except Exception as e:

        return jsonify({
            "status":"error",
            "message":str(e)
        })


# ==========================
# REJECT PASS REQUEST
# ==========================

@app.route("/reject-pass", methods=["POST"])
@login_required("admin")
def reject_pass():

    try:

        data = request.get_json()

        requestId = str(data.get("requestId"))

        req = pd.read_excel(PASS_REQUEST_FILE)

        index = req[
            req["requestId"].astype(str) == requestId
        ].index

        if len(index) == 0:

            return jsonify({

                "status":"error",

                "message":"Request Not Found"

            })

        req.loc[index, "status"] = "Rejected"

        req.loc[index, "approvedBy"] = "Admin"

        req.loc[index, "approvedDate"] = datetime.now().strftime("%d-%m-%Y")

        req.to_excel(
            PASS_REQUEST_FILE,
            index=False
        )

        return jsonify({

            "status":"success",

            "message":"Request Rejected"

        })

    except Exception as e:

        return jsonify({

            "status":"error",

            "message":str(e)

        })
    


# ==========================
# DASHBOARD SUMMARY
# ==========================

@app.route("/dashboard-summary")
@login_required("admin")
def dashboard_summary():

    try:

        emp = load_df()

        totalEmployees = len(emp)

        totalRoutes = emp["route"].fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique()

        pendingRequests = 0
        activePasses = 0

        if os.path.exists(PASS_REQUEST_FILE):

            req = pd.read_excel(PASS_REQUEST_FILE)

            pendingRequests = len(
                req[req["status"] == "Pending"]
            )

        if os.path.exists(PASS_FILE):

            passes = pd.read_excel(PASS_FILE)

            activePasses = len(passes)

        return jsonify({

            "totalEmployees": totalEmployees,

            "totalRoutes": totalRoutes,

            "pendingRequests": pendingRequests,

            "activePasses": activePasses

        })

    except Exception as e:

        return jsonify({

            "status":"error",

            "message":str(e)

        })


# ==========================
# EMPLOYEE SUMMARY (for employee-list.html)
# ==========================

@app.route("/employee-summary")
@login_required("admin")
def employee_summary():

    try:

        emp = load_df()

        totalEmployees = len(emp)

        totalDepartments = emp["department"].fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique()

        totalRoutes = emp["route"].fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique()

        return jsonify({

            "totalEmployees": totalEmployees,

            "totalDepartments": totalDepartments,

            "totalRoutes": totalRoutes

        })

    except Exception as e:

        return jsonify({

            "status":"error",

            "message":str(e)

        })



# ==========================
# GET ALL PASS REQUESTS
# ==========================

@app.route("/admin/pass-requests", methods=["GET"])
@login_required("admin")
def admin_pass_requests():

    try:

        if not os.path.exists(PASS_REQUEST_FILE):

            return jsonify([])

        df = pd.read_excel(PASS_REQUEST_FILE)

        df = df.fillna("")

        return jsonify(
            df.to_dict(orient="records")
        )

    except Exception as e:

        return jsonify({
            "status":"error",
            "message":str(e)
        })
    

    # ==========================
# GENERATED PASSES
# ==========================

@app.route("/generated-passes")
@login_required("admin")
def generated_passes():

    try:

        if not os.path.exists(PASS_FILE):

            return jsonify([])

        passes = pd.read_excel(PASS_FILE)

        return jsonify(

            passes.fillna("").to_dict(
                orient="records"
            )

        )

    except Exception as e:

        return jsonify({

            "status":"error",

            "message":str(e)

        })
    


# ==========================
# GET ALL EMPLOYEES API
# Admin Dashboard → Employee List
# ==========================

@app.route("/employees-json")
@login_required("admin")
def employees_json():

    try:

        # employees.xlsx मधील सर्व employees load करा
        df = load_df()

        # रिकामे values "" करा
        df = df.fillna("")

        # password column कधीही API response मध्ये पाठवायचा नाही
        if "password" in df.columns:
            df = df.drop(columns=["password"])

        # JSON मध्ये return करा
        return jsonify(
            df.to_dict(orient="records")
        )

    except Exception as e:

        return jsonify({
            "status": "error",
            "message": str(e)
        })
    


# ==========================
# TEMPORARY PASS SYSTEM
# ==========================

from datetime import timedelta

TEMP_PASS_FILE = "temp_pass_requests.xlsx"


def create_temp_pass_file():
    if not os.path.exists(TEMP_PASS_FILE):
        df = pd.DataFrame(columns=[
            "requestId", "employeeCode", "employeeName", "department",
            "mobile", "reason", "pickupLocation", "dropLocation",
            "travelDateTime", "requestDate", "status",
            "tempPassId", "validUntil"
        ])
        df.to_excel(TEMP_PASS_FILE, index=False)


create_temp_pass_file()


# ---- Employee: submit a temporary pass request ----
@app.route("/request-temp-pass", methods=["POST"])
@login_required("employee")
def request_temp_pass():

    try:
        data = request.get_json()

        # Always use the logged-in session's own employeeCode, never a
        # client-submitted value.
        employeeCode = str(session.get("employeeCode", "")).strip()

        if employeeCode == "":
            return jsonify({"status": "error", "message": "Employee Code Required"})

        df = pd.read_excel(TEMP_PASS_FILE)

        request_id = "TMP" + str(int(datetime.now().timestamp()))

        new_row = {
            "requestId": request_id,
            "employeeCode": employeeCode,
            "employeeName": data.get("employeeName", ""),
            "department": data.get("department", ""),
            "mobile": data.get("mobile", ""),
            "reason": data.get("reason", ""),
            "pickupLocation": data.get("pickupLocation", ""),
            "dropLocation": data.get("dropLocation", "Bajaj Chakan Plant 1"),
            "travelDateTime": data.get("travelDateTime", ""),
            "requestDate": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "status": "Pending",
            "tempPassId": "",
            "validUntil": ""
        }

        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        df.to_excel(TEMP_PASS_FILE, index=False)

        return jsonify({
            "status": "success",
            "message": "Temporary Pass Request Submitted Successfully",
            "requestId": request_id
        })

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


# ---- Admin: view all temporary pass requests ----
@app.route("/admin/temp-pass-requests", methods=["GET"])
@login_required("admin")
def admin_temp_pass_requests():

    try:
        df = pd.read_excel(TEMP_PASS_FILE)
        df = df.fillna("")
        return jsonify(df.to_dict(orient="records"))

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


# ---- Admin: approve a temporary pass request ----
@app.route("/approve-temp-pass", methods=["POST"])
@login_required("admin")
def approve_temp_pass():

    try:
        data = request.get_json()
        request_id = str(data.get("requestId", ""))

        df = pd.read_excel(TEMP_PASS_FILE)
        match = df["requestId"].astype(str) == request_id

        if not match.any():
            return jsonify({"status": "error", "message": "Request Not Found"})

        temp_pass_id = "TP-" + str(int(datetime.now().timestamp()))
        valid_until = (datetime.now() + timedelta(hours=12)).strftime("%Y-%m-%d %H:%M")

        df.loc[match, "status"] = "Approved"
        df.loc[match, "tempPassId"] = temp_pass_id
        df.loc[match, "validUntil"] = valid_until

        df.to_excel(TEMP_PASS_FILE, index=False)

        return jsonify({
            "status": "success",
            "message": "Temporary Pass Approved"
        })

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


# ---- Admin: reject a temporary pass request ----
@app.route("/reject-temp-pass", methods=["POST"])
@login_required("admin")
def reject_temp_pass():

    try:
        data = request.get_json()
        request_id = str(data.get("requestId", ""))

        df = pd.read_excel(TEMP_PASS_FILE)
        match = df["requestId"].astype(str) == request_id

        if not match.any():
            return jsonify({"status": "error", "message": "Request Not Found"})

        df.loc[match, "status"] = "Rejected"
        df.to_excel(TEMP_PASS_FILE, index=False)

        return jsonify({
            "status": "success",
            "message": "Temporary Pass Rejected"
        })

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


# ---- Employee: view / download their approved temporary pass ----
@app.route("/my-temp-pass", defaults={"employeeCode": None}, methods=["GET"])
@app.route("/my-temp-pass/<employeeCode>", methods=["GET"])
@login_required("employee")
def my_temp_pass(employeeCode):

    # Same fix as /my-pass: always use the session's own employeeCode,
    # never the URL/client value, so a stale/mismatched client value
    # can't wrongly trigger "Not Authorized" for a validly logged-in user.
    employeeCode = session.get("employeeCode")

    try:
        df = pd.read_excel(TEMP_PASS_FILE)

        user_passes = df[
            (df["employeeCode"].astype(str) == str(employeeCode)) &
            (df["status"] == "Approved")
        ]

        if user_passes.empty:
            return jsonify({
                "status": "error",
                "message": "No Approved Temporary Pass Found"
            })

        latest = user_passes.iloc[-1]

        valid_until_str = str(latest.get("validUntil", ""))
        is_expired = False

        try:
            vu = datetime.strptime(valid_until_str, "%Y-%m-%d %H:%M")
            if datetime.now() > vu:
                is_expired = True
        except Exception:
            pass

        return jsonify({
            "status": "success",
            "employeeCode": str(latest["employeeCode"]),
            "employeeName": str(latest["employeeName"]),
            "department": str(latest["department"]),
            "mobile": str(latest["mobile"]),
            "reason": str(latest["reason"]),
            "pickupLocation": str(latest["pickupLocation"]),
            "dropLocation": str(latest["dropLocation"]),
            "travelDateTime": str(latest["travelDateTime"]),
            "tempPassId": str(latest["tempPassId"]),
            "validUntil": valid_until_str,
            "isExpired": is_expired
        })

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


# ==========================
# OT REQUEST
# ==========================


<<<<<<< HEAD


=======
>>>>>>> ca5c43d71517f22e0c26b43691a85c6202fed1b7

OT_FILE = "ot_requests.xlsx"

OT_COLUMNS = [
    "Entry No",
    "Date",
    "Time",
    "Submitted By",
    "Department",
    "Shift",
    "Emergency",
    "Manpower",
    "Transport",
    "Total 2 Hours",
    "Total 3 Hours"
]

COL_WIDTHS = [20, 12, 12, 18, 18, 14, 12, 45, 45, 12, 12]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="00253F", end_color="00253F", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN = Side(style="thin", color="DCE3EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

EMERGENCY_FILL = PatternFill(start_color="FFF7E0", end_color="FFF7E0", fill_type="solid")


def format_manpower(manpower):
    lines = []
    for p in manpower:
        provider = p.get("provider", "Unknown")
        two = p.get("twoHours", 0)
        three = p.get("threeHours", 0)
        if two or three:
            lines.append(f"{provider}: 2Hrs={two}, 3Hrs={three}")
    return "\n".join(lines) if lines else "-"


def format_transport(transport):
    lines = []
    for t in transport:
        route = t.get("route", "Unknown Route")
        two = t.get("twoHours", 0)
        three = t.get("threeHours", 0)
        if two or three:
            lines.append(f"{route}: 2Hrs={two}, 3Hrs={three}")
    return "\n".join(lines) if lines else "-"


def style_header(ws):
    for col_idx, col_name in enumerate(OT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    for i, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def create_ot_file():
    if not os.path.exists(OT_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = datetime.now().strftime("%d-%m-%Y")
        style_header(ws)
        wb.save(OT_FILE)


def append_row_to_sheet(sheet_name, row_values, is_emergency=False):
    wb = load_workbook(OT_FILE)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        style_header(ws)
    else:
        ws = wb[sheet_name]

    next_row = ws.max_row + 1

    for col_idx, value in enumerate(row_values, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.border = BORDER

        if col_idx in (8, 9):
            cell.alignment = BODY_ALIGN
        else:
            cell.alignment = CENTER_ALIGN

        if is_emergency:
            cell.fill = EMERGENCY_FILL

    line_count = max(
        str(row_values[7]).count("\n") + 1,
        str(row_values[8]).count("\n") + 1,
        1
    )
    ws.row_dimensions[next_row].height = 15 * line_count

    wb.save(OT_FILE)


create_ot_file()


# ---- SINGLE route only (removed the duplicate/debug placeholder) ----
@app.route("/save-ot-request", methods=["POST"])
def save_ot_request():
    try:
        data = request.get_json()

        if not data:
            return jsonify({"status": "error", "message": "Invalid Data"})

        manpower = data.get("manpower", [])
        transport = data.get("transport", [])
        emergency = bool(data.get("emergency"))

        total_2h = sum(int(p.get("twoHours", 0) or 0) for p in manpower) + \
                   sum(int(t.get("twoHours", 0) or 0) for t in transport)

        total_3h = sum(int(p.get("threeHours", 0) or 0) for p in manpower) + \
                   sum(int(t.get("threeHours", 0) or 0) for t in transport)

        create_ot_file()

        now = datetime.now()
        sheet_name = now.strftime("%d-%m-%Y")

        row_values = [
            "OT" + now.strftime("%Y%m%d%H%M%S"),
            now.strftime("%d-%m-%Y"),
            now.strftime("%I:%M %p"),
            data.get("submittedBy", ""),
            data.get("department", ""),
            data.get("shift", ""),
            "Yes" if emergency else "No",
            format_manpower(manpower),
            format_transport(transport),
            total_2h,
            total_3h
        ]

        append_row_to_sheet(sheet_name, row_values, is_emergency=emergency)

        saved_path = os.path.abspath(OT_FILE)
        print("SAVED TO:", saved_path)

        return jsonify({
            "status": "success",
            "message": f"Data Saved Successfully at {saved_path}"
        })

    except Exception as e:
        print("SAVE ERROR:", e)
        return jsonify({
            "status": "error",
            "message": str(e)
        })        
# ==========================
# RUN SERVER
# ==========================

if __name__ == "__main__":
    # debug mode should NEVER be on when the site is public (it leaks
    # internal code/data on errors). Set FLASK_DEBUG=True locally if needed.
    app.run(debug=os.environ.get("FLASK_DEBUG", "False") == "True")